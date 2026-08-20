from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from parser.normalize import canonicalize_url


@dataclass(frozen=True)
class HirifyApplicationRow:
    job_title: str
    job_url: str
    company: str
    date_applied: date | None
    stage: str
    feedback: str
    comment: str
    source: str
    recruiter_contact: str
    expected_salary: str
    work_type: str
    created_at: str
    updated_at: str

    @property
    def canonical_job_url(self) -> str:
        return canonicalize_url(self.job_url)

    @property
    def preferred_url(self) -> str:
        contact = (self.recruiter_contact or "").strip()
        if contact.lower().startswith(("http://", "https://")):
            return contact
        return (self.job_url or "").strip()

    @property
    def fingerprint(self) -> str:
        url = self.canonical_job_url or canonicalize_url(self.preferred_url) or self.job_title
        stage = (self.stage or "").strip().lower()
        stamp = (self.updated_at or self.created_at or "").strip()
        if not stamp and self.date_applied:
            stamp = self.date_applied.isoformat()
        return f"{url}|{stage}|{stamp}"


@dataclass(frozen=True)
class StagePlan:
    status: str | None
    close_reason: str | None = None
    closed_stage: str | None = None
    note_only: bool = False


_HEADER_ALIASES = {
    "job title": "job_title",
    "job url": "job_url",
    "company": "company",
    "date applied": "date_applied",
    "stage": "stage",
    "feedback": "feedback",
    "comment": "comment",
    "source": "source",
    "recruiter contact": "recruiter_contact",
    "expected salary": "expected_salary",
    "work type": "work_type",
    "created at": "created_at",
    "updated at": "updated_at",
}


def map_hirify_stage(stage: str) -> StagePlan:
    key = (stage or "").strip().lower()
    if key in {"applied", "viewed", "no response"}:
        return StagePlan(status="Applied")
    if key == "hr interview":
        return StagePlan(status="Interview")
    if key in {"technical interview", "test task"}:
        return StagePlan(status="Interview")
    if key == "final interview":
        return StagePlan(status="Interview")
    if key == "offer":
        return StagePlan(status="Offer")
    if key == "rejected":
        return StagePlan(status="Archived", close_reason="Rejected HR")
    return StagePlan(status="Applied")


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _cell_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def parse_applications_xlsx(path: Path | str) -> list[HirifyApplicationRow]:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        index: dict[str, int] = {}
        for idx, raw in enumerate(header_row):
            alias = _HEADER_ALIASES.get(_cell_str(raw).lower())
            if alias:
                index[alias] = idx

        required = ("job_title", "company", "stage")
        if any(name not in index for name in required):
            raise ValueError(
                "Hirify export missing required columns "
                f"(need Job Title, Company, Stage); got {list(header_row)}"
            )

        out: list[HirifyApplicationRow] = []
        for raw in rows_iter:
            if raw is None or all(cell is None or str(cell).strip() == "" for cell in raw):
                continue

            def get(name: str) -> Any:
                pos = index.get(name)
                if pos is None or pos >= len(raw):
                    return None
                return raw[pos]

            title = _cell_str(get("job_title"))
            company = _cell_str(get("company"))
            if not title and not company:
                continue
            out.append(
                HirifyApplicationRow(
                    job_title=title,
                    job_url=_cell_str(get("job_url")),
                    company=company,
                    date_applied=_cell_date(get("date_applied")),
                    stage=_cell_str(get("stage")),
                    feedback=_cell_str(get("feedback")),
                    comment=_cell_str(get("comment")),
                    source=_cell_str(get("source")),
                    recruiter_contact=_cell_str(get("recruiter_contact")),
                    expected_salary=_cell_str(get("expected_salary")),
                    work_type=_cell_str(get("work_type")),
                    created_at=_cell_str(get("created_at")),
                    updated_at=_cell_str(get("updated_at")),
                )
            )
        return out
    finally:
        workbook.close()


def latest_applications_xlsx(downloads: Path | None = None) -> Path | None:
    root = downloads or Path.home() / "Downloads"
    if not root.is_dir():
        return None
    files = sorted(
        root.glob("my_applications_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return files[0] if files else None
